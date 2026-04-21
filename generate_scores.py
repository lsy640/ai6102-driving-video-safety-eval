"""为三个组员各生成 50 个随机视频编号（00-99），导出 Excel 供打分。"""

import random
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

TOTAL_VIDEOS = 100
SAMPLE_SIZE = 50
MEMBERS = ["Jiang Xinshuo", "Liu Shiyuan", "Liu Yunhao"]

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
header_font = Font(bold=True, size=12)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font_white = Font(bold=True, size=12, color="FFFFFF")
center = Alignment(horizontal="center", vertical="center")

# 保证 100 个视频都被覆盖：先把每个视频至少分给一人，再补齐到 50
all_videos = list(range(TOTAL_VIDEOS))
random.shuffle(all_videos)
assignments = {m: set() for m in MEMBERS}

# 轮流分配，确保每个视频至少被一人覆盖
for idx, vid in enumerate(all_videos):
    assignments[MEMBERS[idx % len(MEMBERS)]].add(vid)

# 每人补齐到 50
for member in MEMBERS:
    remaining = list(set(all_videos) - assignments[member])
    random.shuffle(remaining)
    while len(assignments[member]) < SAMPLE_SIZE:
        assignments[member].add(remaining.pop())

for member in MEMBERS:
    videos = sorted(assignments[member])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "视频打分"

    # 列宽
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 20

    # 表头
    for col, title in enumerate(["视频编号", "评分"], 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    # 数据行
    for row_idx, vid in enumerate(videos, 2):
        c1 = ws.cell(row=row_idx, column=1, value=f"{vid:02d}")
        c1.alignment = center
        c1.border = thin_border

        c2 = ws.cell(row=row_idx, column=2, value="")
        c2.alignment = center
        c2.border = thin_border

    filename = f"{member}_视频打分.xlsx"
    wb.save(filename)
    print(f"{filename}: 视频编号 = {[f'{v:02d}' for v in videos]}")

print("\n完成！共生成 3 个 Excel 文件。")
