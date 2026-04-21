"""读取 auto_evaluation.json，生成 Excel 汇总表。"""

import json
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

with open("auto_evaluation.json", "r") as f:
    data = json.load(f)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "自动评估汇总"

thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, size=11, color="FFFFFF")
safe_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
unsafe_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
center = Alignment(horizontal="center", vertical="center")
wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

headers = ["视频编号", "安全/不安全", "语义级", "逻辑级", "决策级", "问题描述"]
col_widths = [12, 14, 10, 10, 10, 80]

for col, (title, width) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=col, value=title)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = thin_border
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

for row_idx, item in enumerate(data, 2):
    vid = item["video_id"]
    safe = item["safe"]
    errors = item["error_types"]
    desc = item["description"]

    c1 = ws.cell(row=row_idx, column=1, value=vid)
    c1.alignment = center
    c1.border = thin_border

    c2 = ws.cell(row=row_idx, column=2, value="Safe" if safe else "Unsafe")
    c2.alignment = center
    c2.border = thin_border
    c2.fill = safe_fill if safe else unsafe_fill

    for col, etype in zip([3, 4, 5], ["semantic", "logical", "decision"]):
        c = ws.cell(row=row_idx, column=col, value="Y" if etype in errors else "")
        c.alignment = center
        c.border = thin_border
        if etype in errors:
            c.fill = unsafe_fill

    c6 = ws.cell(row=row_idx, column=6, value=desc)
    c6.alignment = wrap
    c6.border = thin_border

# 统计行
stats_row = len(data) + 3
unsafe_count = sum(1 for d in data if not d["safe"])
semantic_count = sum(1 for d in data if "semantic" in d["error_types"])
logical_count = sum(1 for d in data if "logical" in d["error_types"])
decision_count = sum(1 for d in data if "decision" in d["error_types"])

ws.cell(row=stats_row, column=1, value="统计").font = Font(bold=True, size=12)
ws.cell(row=stats_row + 1, column=1, value="总视频数")
ws.cell(row=stats_row + 1, column=2, value=len(data))
ws.cell(row=stats_row + 2, column=1, value="不安全视频")
ws.cell(row=stats_row + 2, column=2, value=unsafe_count)
ws.cell(row=stats_row + 3, column=1, value="不安全比例")
ws.cell(row=stats_row + 3, column=2, value=f"{unsafe_count/len(data)*100:.0f}%")
ws.cell(row=stats_row + 4, column=1, value="语义级错误")
ws.cell(row=stats_row + 4, column=2, value=semantic_count)
ws.cell(row=stats_row + 5, column=1, value="逻辑级错误")
ws.cell(row=stats_row + 5, column=2, value=logical_count)
ws.cell(row=stats_row + 6, column=1, value="决策级错误")
ws.cell(row=stats_row + 6, column=2, value=decision_count)

for r in range(stats_row, stats_row + 7):
    for c in range(1, 3):
        cell = ws.cell(row=r, column=c)
        cell.border = thin_border
        cell.alignment = center

wb.save("自动评估汇总.xlsx")
print(f"生成完成: 自动评估汇总.xlsx")
print(f"总视频: {len(data)}, 不安全: {unsafe_count} ({unsafe_count/len(data)*100:.0f}%)")
print(f"语义级: {semantic_count}, 逻辑级: {logical_count}, 决策级: {decision_count}")
